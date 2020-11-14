import openpyxl
wb = openpyxl.Workbook()
wa = wb.active
# この下にcell指定して書きたいものを入れていく
wa['A1'] = 10
wa['B1'] = "hello"
wa['C1'] = "world"
wa.cell(row=5, column=3, value='こんばんは')

for i in range(3,8):
    for j in range(3,8):
        wa.cell(row=i, column=j).value = 'yeah'

wa = wb.create_sheet(title = "new_sheet")
for i in range(3,12):
    for j in range(3,8):
        wa.cell(row=i, column=j).value = 'check'


wb.save(r'C:\Users\81703\Desktop\sample.xlsx')
# /c/Users/81703/Desktop/python_challenge
print("done")
